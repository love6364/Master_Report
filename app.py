# # app.py
# # ==========================================================
# # SIMPLE FINAL VERSION
# # ONLY WORK:
# # Upload Grid Report + Master Stock File
# # Copy ONLY NUMBERS from Grid -> Master Stock
# # Match Correct Shape
# # Match Correct Size Group
# # Do NOT touch formulas / totals / formatting
# # Download Updated Master Stock File
# # ==========================================================

# import streamlit as st
# from openpyxl import load_workbook
# from io import BytesIO

# # --------------------------------------------------
# # PAGE
# # --------------------------------------------------
# st.set_page_config(page_title="Grid To Master Stock", layout="wide")
# st.title("📊 Grid Report → Master Stock")

# # --------------------------------------------------
# # SHAPES
# # --------------------------------------------------
# SHAPES = [
#     "ROUND",
#     "OVAL",
#     "EMERALD",
#     "RADIANT",
#     "PEAR",
#     "PRINCESS",
#     "MARQUISE",
#     "CUSHION MODIFIED",
#     "CUSHION BRILLIANT",
#     "ASSCHER",
#     "HEART"
# ]

# # --------------------------------------------------
# # HELPERS
# # --------------------------------------------------
# def clean(v):
#     if v is None:
#         return ""
#     return str(v).strip().upper()

# def is_number(v):
#     try:
#         float(v)
#         return True
#     except:
#         return False

# def find_shape_rows(ws):
#     found = {}

#     for r in range(1, ws.max_row + 1):
#         val = clean(ws.cell(r, 1).value)

#         if val in SHAPES and val not in found:
#             found[val] = r

#     return found

# def find_total_row(ws, start_row):
#     for r in range(start_row, ws.max_row + 1):
#         for c in range(1, 10):
#             if clean(ws.cell(r, c).value) == "TOTAL":
#                 return r
#     return start_row + 15

# # --------------------------------------------------
# # MAIN COPY LOGIC
# # --------------------------------------------------
# def copy_grid_to_master(ws_grid, ws_master):

#     grid_shapes = find_shape_rows(ws_grid)
#     master_shapes = find_shape_rows(ws_master)

#     updated = []

#     for shape in SHAPES:

#         if shape in grid_shapes and shape in master_shapes:

#             src_start = grid_shapes[shape]
#             src_end   = find_total_row(ws_grid, src_start)

#             tgt_start = master_shapes[shape]

#             rows = src_end - src_start

#             for i in range(rows):

#                 src_row = src_start + i
#                 tgt_row = tgt_start + i

#                 for col in range(1, ws_grid.max_column + 1):

#                     val = ws_grid.cell(src_row, col).value

#                     # only numeric copy
#                     if is_number(val):
#                         ws_master.cell(tgt_row, col).value = val

#             updated.append(shape)

#     return updated

# # --------------------------------------------------
# # PROCESS
# # --------------------------------------------------
# def process(grid_file, master_file):

#     wb_grid = load_workbook(grid_file, data_only=True)
#     wb_master = load_workbook(master_file)

#     ws_grid = wb_grid.active
#     ws_master = wb_master.active

#     updated = copy_grid_to_master(ws_grid, ws_master)

#     output = BytesIO()
#     wb_master.save(output)
#     output.seek(0)

#     return output, updated

# # --------------------------------------------------
# # UI
# # --------------------------------------------------
# grid_file = st.file_uploader("Upload Grid Report", type=["xlsx"])
# master_file = st.file_uploader("Upload Master Stock File", type=["xlsx"])

# if grid_file and master_file:

#     if st.button("🚀 Process"):

#         try:
#             with st.spinner("Working..."):

#                 output, updated = process(grid_file, master_file)

#             st.success("✅ Completed Successfully")

#             st.write("### Updated Shapes")
#             st.write(", ".join(updated))

#             st.download_button(
#                 "📥 Download Updated Master Stock File",
#                 data=output,
#                 file_name="Updated_Master_Stock.xlsx",
#                 mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#             )

#         except Exception as e:
#             st.error(str(e))


# # app.py
# # ==========================================================
# # FIXED VERSION
# # Copy/Paste VALUES ONLY
# # No formulas inserted
# # No #REF issue
# # Download all 3 files
# # # ==========================================================

# import streamlit as st
# from openpyxl import load_workbook
# from io import BytesIO

# st.set_page_config(page_title="Daily Work Automation", layout="wide")
# st.title("📊 Full Daily Work Automation")

# # --------------------------------------------------
# # SESSION
# # --------------------------------------------------
# if "processed" not in st.session_state:
#     st.session_state.processed = False

# # --------------------------------------------------
# # SHAPES
# # --------------------------------------------------
# SHAPES = [
#     "ROUND","OVAL","EMERALD","RADIANT","PEAR",
#     "PRINCESS","MARQUISE","CUSHION MODIFIED",
#     "CUSHION BRILLIANT","ASSCHER","HEART"
# ]

# # --------------------------------------------------
# # HELPERS
# # --------------------------------------------------
# def clean(v):
#     if v is None:
#         return ""
#     return str(v).strip().upper()

# def is_number(v):
#     try:
#         float(v)
#         return True
#     except:
#         return False

# def find_header(ws, text):
#     for row in ws.iter_rows():
#         for cell in row:
#             if clean(cell.value) == clean(text):
#                 return cell.row, cell.column
#     return None, None

# def find_shape_rows(ws):
#     found = {}

#     for r in range(1, ws.max_row + 1):
#         val = clean(ws.cell(r,1).value)

#         if val in SHAPES and val not in found:
#             found[val] = r

#     return found

# def find_total_row(ws, start):
#     for r in range(start, ws.max_row + 1):
#         for c in range(1,10):
#             if clean(ws.cell(r,c).value) == "TOTAL":
#                 return r
#     return start + 15

# # --------------------------------------------------
# # STEP 1 GRID -> MASTER STOCK
# # --------------------------------------------------
# def copy_grid_to_stock(ws_grid, ws_stock):

#     g = find_shape_rows(ws_grid)
#     s = find_shape_rows(ws_stock)

#     for shape in SHAPES:

#         if shape in g and shape in s:

#             gs = g[shape]
#             ge = find_total_row(ws_grid, gs)

#             ss = s[shape]

#             rows = ge - gs

#             for i in range(rows):

#                 for col in range(1, ws_grid.max_column + 1):

#                     val = ws_grid.cell(gs+i, col).value

#                     if is_number(val):
#                         ws_stock.cell(ss+i, col).value = val

# # --------------------------------------------------
# # COPY VALUE ONLY
# # --------------------------------------------------
# def copy_value_column(ws_source, source_header, ws_target, target_header):

#     sr, sc = find_header(ws_source, source_header)
#     tr, tc = find_header(ws_target, target_header)

#     if not sr or not tr:
#         return

#     r = sr + 1
#     t = tr + 1

#     while r <= ws_source.max_row:

#         val = ws_source.cell(r, sc).value

#         if r > sr + 1000:
#             break

#         # IMPORTANT: copy only final values
#         if val is None or clean(val) == "":
#             ws_target.cell(t, tc).value = 0
#         else:
#             ws_target.cell(t, tc).value = val

#         r += 1
#         t += 1

# # --------------------------------------------------
# # PROCESS
# # --------------------------------------------------
# def process(grid_file, stock_file, master_file):

#     # normal files
#     wb_grid = load_workbook(grid_file)
#     wb_stock = load_workbook(stock_file)
#     wb_master = load_workbook(master_file)

#     ws_grid = wb_grid.active
#     ws_stock = wb_stock.active
#     ws_master = wb_master.active

#     # ------------------------------------------------
#     # STEP 1
#     # Grid -> Master Stock
#     # ------------------------------------------------
#     copy_grid_to_stock(ws_grid, ws_stock)

#     # Save updated stock first
#     temp = BytesIO()
#     wb_stock.save(temp)
#     temp.seek(0)

#     # ------------------------------------------------
#     # IMPORTANT FIX
#     # Load SAME updated file values
#     # ------------------------------------------------
#     wb_stock_values = load_workbook(temp, data_only=False)

#     # Sheet2
#     ws_sheet2 = wb_stock_values.worksheets[1]

#     # ------------------------------------------------
#     # STEP 2
#     # INHAND -> Available
#     # MAX PCS -> Grid
#     # ------------------------------------------------
#     copy_value_column(ws_sheet2, "INHAND", ws_master, "AVAILABLE")
#     copy_value_column(ws_sheet2, "MAX PCS", ws_grid, "GRID")

#     # ------------------------------------------------
#     # SAVE
#     # ------------------------------------------------
#     out1 = BytesIO()
#     out2 = BytesIO()
#     out3 = BytesIO()

#     wb_grid.save(out1)
#     wb_stock.save(out2)
#     wb_master.save(out3)

#     out1.seek(0)
#     out2.seek(0)
#     out3.seek(0)

#     return out1, out2, out3

# # --------------------------------------------------
# # UI
# # --------------------------------------------------
# grid_file = st.file_uploader("Upload Grid Report", type=["xlsx"])
# stock_file = st.file_uploader("Upload Master Stock File", type=["xlsx"])
# master_file = st.file_uploader("Upload Master File", type=["xlsx"])

# if grid_file and stock_file and master_file:

#     if st.button("🚀 Process All Files"):

#         try:
#             with st.spinner("Processing..."):

#                 g, s, m = process(grid_file, stock_file, master_file)

#                 st.session_state.grid = g.getvalue()
#                 st.session_state.stock = s.getvalue()
#                 st.session_state.master = m.getvalue()

#                 st.session_state.processed = True

#             st.success("✅ Completed Successfully")

#         except Exception as e:
#             st.error(str(e))

# # --------------------------------------------------
# # DOWNLOADS
# # --------------------------------------------------
# if st.session_state.processed:

#     st.download_button(
#         "📥 Download Grid Report",
#         data=st.session_state.grid,
#         file_name="Updated_Grid_Report.xlsx"
#     )

#     st.download_button(
#         "📥 Download Master Stock",
#         data=st.session_state.stock,
#         file_name="Updated_Master_Stock.xlsx"
#     )

#     st.download_button(
#         "📥 Download Master File",
#         data=st.session_state.master,
#         file_name="Updated_Master_File.xlsx"
#     )


# # app.py - Daily Work Automation
# import streamlit as st
# from openpyxl import load_workbook
# from io import BytesIO

# st.set_page_config(page_title="Diamond Work Automation", layout="wide", page_icon="💎")

# st.markdown("""
# <style>
# @import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600&family=DM+Mono&display=swap');
# html, body, [class*="css"] { font-family: 'DM Sans', sans-serif; }
# .stApp { background: #0f1117; color: #e8eaf0; }
# header[data-testid="stHeader"] { background: transparent; }
# .upload-label { font-size: 0.78rem; font-weight: 500; letter-spacing: 1.2px; text-transform: uppercase; color: #4f6ef7; margin-bottom: 0.5rem; }
# .steps-panel { background: #1a1d27; border: 1px solid #2a2d3a; border-radius: 14px; padding: 1.4rem 1.6rem; }
# .steps-panel h4 { color: #7c8394; font-size: 0.72rem; letter-spacing: 1.4px; text-transform: uppercase; margin-bottom: 1rem; font-weight: 500; }
# .step-row { display: flex; align-items: flex-start; gap: 0.9rem; margin-bottom: 0.85rem; }
# .step-num { background: #4f6ef7; color: white; border-radius: 50%; width: 22px; height: 22px; font-size: 0.7rem; font-weight: 600; display: flex; align-items: center; justify-content: center; flex-shrink: 0; margin-top: 1px; }
# .step-text { font-size: 0.88rem; color: #c0c4d0; line-height: 1.5; }
# .step-text span { color: #ffffff; font-weight: 500; }
# .step-arrow { color: #4f6ef7; font-size: 0.8rem; }
# div[data-testid="stButton"] > button { background: linear-gradient(135deg, #4f6ef7 0%, #7c4ff7 100%); color: white; border: none; border-radius: 10px; padding: 0.75rem 2rem; font-family: 'DM Sans', sans-serif; font-size: 0.95rem; font-weight: 500; letter-spacing: 0.3px; width: 100%; cursor: pointer; box-shadow: 0 4px 20px rgba(79, 110, 247, 0.3); transition: opacity 0.2s, transform 0.1s; }
# div[data-testid="stButton"] > button:hover { opacity: 0.9; transform: translateY(-1px); }
# div[data-testid="stDownloadButton"] > button { background: #1a1d27; color: #e8eaf0; border: 1px solid #2a2d3a; border-radius: 10px; padding: 0.65rem 1.2rem; font-family: 'DM Sans', sans-serif; font-size: 0.88rem; font-weight: 400; width: 100%; transition: border-color 0.2s, background 0.2s; }
# div[data-testid="stDownloadButton"] > button:hover { border-color: #4f6ef7; background: #1f2235; color: #ffffff; }
# div[data-testid="stFileUploader"] section { border: 1.5px dashed #2a2d3a; border-radius: 10px; background: #12151e; padding: 0.8rem; transition: border-color 0.2s; }
# div[data-testid="stFileUploader"] section:hover { border-color: #4f6ef7; }
# hr { border-color: #2a2d3a; margin: 1.5rem 0; }
# .dl-header { font-size: 0.75rem; font-weight: 500; letter-spacing: 1.3px; text-transform: uppercase; color: #7c8394; margin-bottom: 0.8rem; margin-top: 1.2rem; }
# .hero { text-align: center; padding: 2.5rem 0 1.5rem 0; }
# .hero h1 { font-size: 2.4rem; font-weight: 600; letter-spacing: -0.5px; color: #ffffff; margin-bottom: 0.3rem; }
# .hero p { color: #7c8394; font-size: 1rem; font-weight: 300; }
# </style>
# """, unsafe_allow_html=True)

# # --------------------------------------------------
# # SESSION
# # --------------------------------------------------
# if "processed" not in st.session_state:
#     st.session_state.processed = False

# # --------------------------------------------------
# # SHAPES
# # --------------------------------------------------
# SHAPES = [
#     "ROUND","OVAL","EMERALD","RADIANT","PEAR",
#     "PRINCESS","MARQUISE","CUSHION MODIFIED",
#     "CUSHION BRILLIANT","ASSCHER","HEART","SQ RADIANT"
# ]

# # --------------------------------------------------
# # HELPERS
# # --------------------------------------------------
# def clean(v):
#     if v is None:
#         return ""
#     return str(v).strip().upper()

# def is_number(v):
#     try:
#         float(v)
#         return True
#     except:
#         return False

# def find_header(ws, text):
#     for row in ws.iter_rows():
#         for cell in row:
#             if clean(cell.value) == clean(text):
#                 return cell.row, cell.column
#     return None, None

# def find_shape_rows(ws):
#     found = {}
#     for r in range(1, ws.max_row + 1):
#         val = clean(ws.cell(r, 1).value)
#         if val in SHAPES and val not in found:
#             found[val] = r
#     return found

# def find_total_row(ws, start):
#     for r in range(start, ws.max_row + 1):
#         for c in range(1, 10):
#             if clean(ws.cell(r, c).value) == "TOTAL":
#                 return r
#     return start + 15

# # --------------------------------------------------
# # STEP 1 + MISSING SHAPE BLANK:
# # Grid -> Master Stock shape-wise numeric copy
# # If shape is IN Master Stock but NOT in Grid -> blank ALL numeric cols for that shape
# # --------------------------------------------------
# def copy_grid_to_stock(ws_grid, ws_stock):
#     g = find_shape_rows(ws_grid)
#     s = find_shape_rows(ws_stock)

#     for shape in SHAPES:
#         if shape not in s:
#             continue  # shape not in master stock at all, skip

#         ss = s[shape]
#         se = find_total_row(ws_stock, ss)
#         rows = se - ss

#         if shape in g:
#             # Shape exists in Grid -> copy numeric values
#             gs = g[shape]
#             ge = find_total_row(ws_grid, gs)
#             grid_rows = ge - gs

#             for i in range(rows):
#                 for col in range(1, ws_grid.max_column + 1):
#                     if i < grid_rows:
#                         val = ws_grid.cell(gs + i, col).value
#                         if is_number(val):
#                             ws_stock.cell(ss + i, col).value = val
#                     # if grid has fewer rows, don't touch remaining stock rows
#         else:
#             # ✅ Shape MISSING from Grid -> blank ALL numeric columns for this shape in Master Stock
#             for i in range(rows):
#                 for col in range(4, ws_stock.max_column + 1):  # col 1-3 = Shape/Color/Clarity labels, keep those
#                     val = ws_stock.cell(ss + i, col).value
#                     if is_number(val):
#                         ws_stock.cell(ss + i, col).value = None

# # --------------------------------------------------
# # STEP 2: MAX PCS from Grid -> Master Stock (shape-wise, blank if missing)
# # --------------------------------------------------
# def copy_max_pcs_grid_to_stock(ws_grid, ws_stock):
#     grid_hr, grid_hc = find_header(ws_grid, "MAX PCS")
#     stock_hr, stock_hc = find_header(ws_stock, "MAX PCS")
#     if not grid_hr or not stock_hr:
#         return

#     g_shapes = find_shape_rows(ws_grid)
#     s_shapes = find_shape_rows(ws_stock)

#     for shape in SHAPES:
#         if shape not in s_shapes:
#             continue

#         ss = s_shapes[shape]
#         se = find_total_row(ws_stock, ss)
#         rows = se - ss

#         if shape in g_shapes:
#             gs = g_shapes[shape]
#             ge = find_total_row(ws_grid, gs)
#             grid_rows = ge - gs
#             for i in range(rows):
#                 if i < grid_rows:
#                     val = ws_grid.cell(gs + i, grid_hc).value
#                     ws_stock.cell(ss + i, stock_hc).value = 0 if (val is None or clean(val) == "") else val
#                 else:
#                     ws_stock.cell(ss + i, stock_hc).value = None
#         else:
#             # Shape missing from Grid -> blank MAX PCS in Master Stock
#             for i in range(rows):
#                 ws_stock.cell(ss + i, stock_hc).value = None

# # --------------------------------------------------
# # COPY VALUE COLUMN
# # --------------------------------------------------
# def copy_value_column(ws_source, source_header, ws_target, target_header):
#     sr, sc = find_header(ws_source, source_header)
#     tr, tc = find_header(ws_target, target_header)
#     if not sr or not tr:
#         return
#     r, t = sr + 1, tr + 1
#     while r <= ws_source.max_row:
#         if r > sr + 1000:
#             break
#         val = ws_source.cell(r, sc).value
#         ws_target.cell(t, tc).value = 0 if (val is None or clean(val) == "") else val
#         r += 1
#         t += 1

# # --------------------------------------------------
# # PROCESS
# # --------------------------------------------------
# def process(grid_file, stock_file, master_file):
#     wb_grid   = load_workbook(grid_file)
#     wb_stock  = load_workbook(stock_file)
#     wb_master = load_workbook(master_file)

#     ws_grid   = wb_grid.active
#     ws_stock  = wb_stock.active
#     ws_master = wb_master.active

#     # STEP 1: Grid -> Master Stock + blank missing shapes
#     copy_grid_to_stock(ws_grid, ws_stock)

#     # STEP 2: MAX PCS Grid -> Master Stock (shape-wise, blanks missing)
#     copy_max_pcs_grid_to_stock(ws_grid, ws_stock)

#     # Save stock, reload to get updated Sheet2
#     temp = BytesIO()
#     wb_stock.save(temp)
#     temp.seek(0)
#     wb_stock_values = load_workbook(temp, data_only=False)
#     ws_sheet2 = wb_stock_values.worksheets[1]

#     # STEP 3: Sheet2 INHAND -> Master File AVAILABLE
#     copy_value_column(ws_sheet2, "INHAND", ws_master, "AVAILABLE")

#     # STEP 4: Sheet2 MAX PCS -> Master File GRID
#     copy_value_column(ws_sheet2, "MAX PCS", ws_master, "GRID")

#     out1, out2, out3 = BytesIO(), BytesIO(), BytesIO()
#     wb_grid.save(out1)
#     wb_stock.save(out2)
#     wb_master.save(out3)
#     out1.seek(0); out2.seek(0); out3.seek(0)
#     return out1, out2, out3

# # --------------------------------------------------
# # UI — HERO
# # --------------------------------------------------
# st.markdown("""
# <div class="hero">
#     <h1>💎 Diamond Work Automation</h1>
#     <p>Upload your three files, process in one click, download updated reports.</p>
# </div>
# """, unsafe_allow_html=True)

# col_left, col_right = st.columns([3, 2], gap="large")

# with col_left:
#     st.markdown('<div class="upload-label">📂 Grid Report</div>', unsafe_allow_html=True)
#     grid_file = st.file_uploader("Grid Report", type=["xlsx"], label_visibility="collapsed")

#     st.markdown('<div class="upload-label" style="margin-top:1rem;">📦 Master Stock File</div>', unsafe_allow_html=True)
#     stock_file = st.file_uploader("Master Stock", type=["xlsx"], label_visibility="collapsed")

#     st.markdown('<div class="upload-label" style="margin-top:1rem;">📋 Master File</div>', unsafe_allow_html=True)
#     master_file = st.file_uploader("Master File", type=["xlsx"], label_visibility="collapsed")

#     st.markdown("<div style='margin-top:1.5rem;'></div>", unsafe_allow_html=True)

#     all_uploaded = grid_file and stock_file and master_file

#     if all_uploaded:
#         if st.button("🚀 Process All Files"):
#             try:
#                 with st.spinner("Processing your files..."):
#                     g, s, m = process(grid_file, stock_file, master_file)
#                     st.session_state.grid   = g.getvalue()
#                     st.session_state.stock  = s.getvalue()
#                     st.session_state.master = m.getvalue()
#                     st.session_state.processed = True
#                 st.success("✅ All files processed successfully!")
#             except Exception as e:
#                 st.error(f"❌ Error: {e}")
#     else:
#         st.button("🚀 Process All Files", disabled=True)
#         st.markdown("<p style='color:#4a4f63; font-size:0.82rem; margin-top:0.4rem;'>Upload all 3 files to enable processing.</p>", unsafe_allow_html=True)

# with col_right:
#     st.markdown("""
#     <div class="steps-panel">
#         <h4>What this does</h4>
#         <div class="step-row">
#             <div class="step-num">1</div>
#             <div class="step-text">
#                 <span>Grid Report</span> → <span>Master Stock</span><br>
#                 <span class="step-arrow">↳</span> Shape-wise numeric copy · <span style="color:#f87171;">blanks entire shape if missing from Grid</span>
#             </div>
#         </div>
#         <div class="step-row">
#             <div class="step-num">2</div>
#             <div class="step-text">
#                 <span>Grid Report MAX PCS</span> → <span>Master Stock MAX PCS</span><br>
#                 <span class="step-arrow">↳</span> Shape-wise · blanks if shape missing in Grid
#             </div>
#         </div>
#         <div class="step-row">
#             <div class="step-num">3</div>
#             <div class="step-text">
#                 <span>Master Stock Sheet2 INHAND</span> → <span>Master File AVAILABLE</span>
#             </div>
#         </div>
#         <div class="step-row">
#             <div class="step-num">4</div>
#             <div class="step-text">
#                 <span>Master Stock Sheet2 MAX PCS</span> → <span>Master File GRID</span>
#             </div>
#         </div>
#     </div>
#     """, unsafe_allow_html=True)

# # --------------------------------------------------
# # DOWNLOADS
# # --------------------------------------------------
# if st.session_state.processed:
#     st.markdown("<hr>", unsafe_allow_html=True)
#     st.markdown('<div class="dl-header">📥 Download Updated Files</div>', unsafe_allow_html=True)

#     d1, d2, d3 = st.columns(3, gap="small")
#     with d1:
#         st.download_button("📊 Grid Report", data=st.session_state.grid, file_name="Updated_Grid_Report.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
#     with d2:
#         st.download_button("📦 Master Stock", data=st.session_state.stock, file_name="Updated_Master_Stock.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
#     with d3:
#         st.download_button("📋 Master File", data=st.session_state.master, file_name="Updated_Master_File.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# import streamlit as st
# import pandas as pd
# import numpy as np
# from io import BytesIO
# import openpyxl
# from openpyxl import load_workbook
# import copy

# # ---------------- PAGE CONFIG ----------------
# st.set_page_config(
#     page_title="DiamondFlow AI",
#     page_icon="💎",
#     layout="wide",
#     initial_sidebar_state="collapsed"
# )

# # ---------------- PREMIUM CSS ----------------
# st.markdown("""
# <style>
# header {visibility:hidden;}
# #MainMenu {visibility:hidden;}
# footer {visibility:hidden;}

# .stApp {
#     background: radial-gradient(circle at top left,#00111f 0%,transparent 35%),
#                 radial-gradient(circle at top right,#1d0038 0%,transparent 35%),
#                 linear-gradient(135deg,#050505,#0c0c0c,#101010,#050505);
#     color: white;
# }

# .block-container {
#     padding-top: 0.6rem !important;
#     padding-bottom: 2rem;
#     max-width: 1450px;
# }

# .main-title {
#     text-align: center;
#     font-size: 2.9rem;
#     font-weight: 900;
#     margin-top: 0px;
#     margin-bottom: 0px;
#     line-height: 1.1;
#     letter-spacing: 1px;
#     background: linear-gradient(90deg,#3da5ff,#7f6bff,#00f0ff,#3da5ff);
#     background-size: 250% 250%;
#     animation: shine 6s linear infinite;
#     -webkit-background-clip: text;
#     -webkit-text-fill-color: transparent;
# }

# .sub-title {
#     text-align: center;
#     color: #d9d9d9;
#     font-size: 14px;
#     margin-top: 6px;
#     margin-bottom: 18px;
#     letter-spacing: 0.4px;
# }

# @keyframes shine {
#     0%   { background-position: 0% 50%; }
#     100% { background-position: 100% 50%; }
# }

# hr {
#     border: none;
#     height: 1px;
#     background: linear-gradient(to right,transparent,#3da5ff,#00f0ff,#7f6bff,transparent);
#     margin: 18px 0;
# }

# .upload-box {
#     background: rgba(255,255,255,0.04);
#     border: 1px solid rgba(255,255,255,0.08);
#     border-radius: 18px;
#     padding: 18px;
#     backdrop-filter: blur(12px);
#     box-shadow: 0 10px 28px rgba(0,0,0,0.30);
#     transition: 0.25s;
#     margin-bottom: 12px;
# }

# .upload-box:hover {
#     transform: translateY(-3px);
#     border: 1px solid rgba(61,165,255,0.45);
# }

# .upload-title {
#     color: #3da5ff;
#     font-size: 17px;
#     font-weight: 800;
# }

# .upload-desc {
#     color: #dddddd;
#     font-size: 13px;
#     margin-top: 4px;
# }

# .step-badge {
#     display: inline-block;
#     background: linear-gradient(135deg,#3da5ff,#7f6bff);
#     color: white;
#     font-size: 11px;
#     font-weight: 800;
#     padding: 3px 10px;
#     border-radius: 20px;
#     margin-bottom: 6px;
#     letter-spacing: 1px;
# }

# .success-box {
#     background: rgba(0,240,100,0.08);
#     border: 1px solid rgba(0,240,100,0.3);
#     border-radius: 14px;
#     padding: 14px 18px;
#     color: #00f064;
#     font-size: 14px;
#     margin: 10px 0;
# }

# .info-box {
#     background: rgba(61,165,255,0.08);
#     border: 1px solid rgba(61,165,255,0.3);
#     border-radius: 14px;
#     padding: 14px 18px;
#     color: #3da5ff;
#     font-size: 13px;
#     margin: 8px 0;
# }

# .warn-box {
#     background: rgba(255,180,0,0.08);
#     border: 1px solid rgba(255,180,0,0.3);
#     border-radius: 14px;
#     padding: 14px 18px;
#     color: #ffb400;
#     font-size: 13px;
#     margin: 8px 0;
# }

# [data-testid="stFileUploader"] {
#     border: 2px dashed rgba(61,165,255,0.35);
#     border-radius: 16px;
#     padding: 12px;
#     background: rgba(255,255,255,0.03);
# }

# [data-testid="stFileUploader"]:hover {
#     border-color: #00f0ff;
# }

# .stButton>button, .stDownloadButton>button {
#     width: 100%;
#     border: none;
#     border-radius: 14px;
#     padding: 14px;
#     font-size: 15px;
#     font-weight: 800;
#     color: white;
#     background: linear-gradient(135deg,#3da5ff,#7f6bff,#00f0ff);
#     background-size: 250% 250%;
#     animation: shine 6s linear infinite;
#     box-shadow: 0 12px 25px rgba(0,0,0,0.30);
# }

# .stButton>button:hover, .stDownloadButton>button:hover {
#     transform: translateY(-2px);
# }

# [data-testid="stMetric"] {
#     background: rgba(255,255,255,0.04);
#     border: 1px solid rgba(255,255,255,0.08);
#     border-radius: 16px;
#     padding: 15px;
# }

# [data-testid="stMetricValue"] {
#     color: #00f0ff;
# }

# [data-testid="stDataFrame"] {
#     border-radius: 16px;
#     overflow: hidden;
#     border: 1px solid rgba(255,255,255,0.08);
# }
# </style>
# """, unsafe_allow_html=True)

# # ---------------- HEADER ----------------
# st.markdown('<div class="main-title">DiamondFlow AI 💎</div>', unsafe_allow_html=True)
# st.markdown('<div class="sub-title">Luxury Diamond Inventory Automation Platform</div>', unsafe_allow_html=True)
# st.markdown("<hr>", unsafe_allow_html=True)


# # ---------------- HELPER: normalize key ----------------
# def make_key(shape, color, clarity, from_size, to_size):
#     """Create a normalized composite key for matching rows."""
#     def clean(v):
#         if v is None:
#             return ""
#         return str(v).strip().upper()
#     return (clean(shape), clean(color), clean(clarity), clean(from_size), clean(to_size))


# # ---------------- STEP 1: Read INHAND from Grid Report ----------------
# def read_grid_inhand(grid_file):
#     """
#     Grid Report → Sheet1
#     Reads SHAPE, COLOR, CLARITY, SIZE (FROM SIZE, TO SIZE), INHAND_PCS columns.
#     Finds INHAND_PCS column header and SIZE/FROM SIZE/TO SIZE dynamically.
#     Returns a dict: key(shape,color,clarity,from_size,to_size) → inhand_value
#     """
#     wb = load_workbook(grid_file, data_only=True)
#     ws = wb.active  # Sheet1

#     # Find header row — look for a row containing 'SHAPE' or 'COLOR'
#     header_row_idx = None
#     headers = {}
#     for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
#         row_upper = [str(c).strip().upper() if c is not None else "" for c in row]
#         if "SHAPE" in row_upper and "COLOR" in row_upper:
#             header_row_idx = row_idx
#             for col_idx, val in enumerate(row_upper):
#                 headers[val] = col_idx
#             break

#     if header_row_idx is None:
#         raise ValueError("Could not find header row in Grid Report (looking for SHAPE, COLOR columns)")

#     # Find required columns
#     def find_col(names):
#         for name in names:
#             if name in headers:
#                 return headers[name]
#         return None

#     shape_col  = find_col(["SHAPE"])
#     color_col  = find_col(["COLOR"])
#     clarity_col= find_col(["CLARITY"])
#     inhand_col = find_col(["INHAND_PCS", "INHAND PCS", "INHAND"])

#     # For size, look for FROM SIZE / TO SIZE or SIZE
#     from_size_col = find_col(["FROM SIZE", "FROM_SIZE", "FROM"])
#     to_size_col   = find_col(["TO SIZE", "TO_SIZE", "TO"])
#     size_col      = find_col(["SIZE"])

#     if inhand_col is None:
#         raise ValueError("Could not find INHAND_PCS column in Grid Report. Found columns: " + str(list(headers.keys())))
#     if shape_col is None or color_col is None or clarity_col is None:
#         raise ValueError("Could not find SHAPE/COLOR/CLARITY columns in Grid Report.")

#     inhand_map = {}
#     all_rows = list(ws.iter_rows(values_only=True))

#     for row in all_rows[header_row_idx:]:  # data starts after header
#         shape   = row[shape_col]   if shape_col  is not None else None
#         color   = row[color_col]   if color_col  is not None else None
#         clarity = row[clarity_col] if clarity_col is not None else None
#         inhand  = row[inhand_col]

#         if not shape or str(shape).strip().upper() in ["", "SHAPE", "TOTAL", "CS_TOTAL"]:
#             continue

#         # Determine from/to size
#         if from_size_col is not None and to_size_col is not None:
#             from_size = row[from_size_col]
#             to_size   = row[to_size_col]
#         elif size_col is not None:
#             size_val = row[size_col]
#             # size_val might be "0.30 - 0.49" or just a number
#             if size_val and "-" in str(size_val):
#                 parts = str(size_val).split("-")
#                 from_size = parts[0].strip()
#                 to_size   = parts[1].strip()
#             else:
#                 from_size = size_val
#                 to_size   = size_val
#         else:
#             from_size = ""
#             to_size   = ""

#         key = make_key(shape, color, clarity, from_size, to_size)
#         try:
#             inhand_val = float(inhand) if inhand is not None else 0
#         except (ValueError, TypeError):
#             inhand_val = 0

#         inhand_map[key] = inhand_val

#     return inhand_map


# # ---------------- STEP 2: Read MAX PCS + INHAND from Master Stock MASTER sheet ----------------
# def read_master_stock(master_stock_file):
#     """
#     Master Stock → MASTER sheet
#     Reads SHAPE, FROM SIZE, TO SIZE, COLOR, CLARITY, MAX PCS, INHAND
#     Returns dict: key → {"max_pcs": val, "inhand": val}
#     """
#     wb = load_workbook(master_stock_file, data_only=True)

#     # Try to find the MASTER sheet
#     sheet_names_upper = {s.upper(): s for s in wb.sheetnames}
#     if "MASTER" in sheet_names_upper:
#         ws = wb[sheet_names_upper["MASTER"]]
#     else:
#         # fallback to first sheet
#         ws = wb.active

#     # Find header row
#     header_row_idx = None
#     headers = {}
#     for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
#         row_upper = [str(c).strip().upper() if c is not None else "" for c in row]
#         if "SHAPE" in row_upper and ("MAX PCS" in row_upper or "MAX_PCS" in row_upper or "INHAND" in row_upper):
#             header_row_idx = row_idx
#             for col_idx, val in enumerate(row_upper):
#                 if val:
#                     headers[val] = col_idx
#             break

#     if header_row_idx is None:
#         # Try row 1
#         row = list(ws.iter_rows(values_only=True))[0]
#         row_upper = [str(c).strip().upper() if c is not None else "" for c in row]
#         header_row_idx = 1
#         for col_idx, val in enumerate(row_upper):
#             if val:
#                 headers[val] = col_idx

#     def find_col(names):
#         for name in names:
#             if name in headers:
#                 return headers[name]
#         return None

#     shape_col    = find_col(["SHAPE"])
#     color_col    = find_col(["COLOR"])
#     clarity_col  = find_col(["CLARITY"])
#     from_size_col= find_col(["FROM SIZE", "FROM_SIZE", "FROM"])
#     to_size_col  = find_col(["TO SIZE", "TO_SIZE", "TO"])
#     size_col     = find_col(["SIZE"])
#     max_pcs_col  = find_col(["MAX PCS", "MAX_PCS", "MAXPCS"])
#     inhand_col   = find_col(["INHAND", "INHAND_PCS", "IN HAND"])

#     data_map = {}
#     all_rows = list(ws.iter_rows(values_only=True))

#     for row in all_rows[header_row_idx:]:
#         shape = row[shape_col] if shape_col is not None else None
#         if not shape or str(shape).strip().upper() in ["", "SHAPE", "TOTAL"]:
#             continue

#         color   = row[color_col]   if color_col   is not None else None
#         clarity = row[clarity_col] if clarity_col is not None else None

#         if from_size_col is not None and to_size_col is not None:
#             from_size = row[from_size_col]
#             to_size   = row[to_size_col]
#         elif size_col is not None:
#             size_val = row[size_col]
#             if size_val and "-" in str(size_val):
#                 parts = str(size_val).split("-")
#                 from_size = parts[0].strip()
#                 to_size   = parts[1].strip()
#             else:
#                 from_size = size_val
#                 to_size   = size_val
#         else:
#             from_size = ""
#             to_size   = ""

#         max_pcs = row[max_pcs_col] if max_pcs_col is not None else None
#         inhand  = row[inhand_col]  if inhand_col  is not None else None

#         try:
#             max_pcs_val = float(max_pcs) if max_pcs is not None else 0
#         except (ValueError, TypeError):
#             max_pcs_val = 0

#         try:
#             inhand_val = float(inhand) if inhand is not None else 0
#         except (ValueError, TypeError):
#             inhand_val = 0

#         key = make_key(shape, color, clarity, from_size, to_size)
#         data_map[key] = {"max_pcs": max_pcs_val, "inhand": inhand_val}

#     return data_map


# # ---------------- STEP 3: Update Master File ----------------
# def update_master_file(master_file_bytes, inhand_map_grid, master_stock_map):
#     """
#     Master File:
#       - Columns: Shape(A), From Size(B), To Size(C), Clarity(D), Color(E), Grid(F), Available(G)
#       - Step A: Fill Grid (col F) with INHAND from Grid Report
#       - Step B: Fill Grid (col F) with Max PCS from Master Stock (overwrite if Grid still empty/0,
#                 OR if master_stock has value — per requirement: max pcs → Grid column)
#       - Step C: Fill Available (col G) with INHAND from Master Stock MASTER sheet
#       - Blanks filled with 0
#     Returns modified workbook as bytes.
#     """
#     wb = load_workbook(BytesIO(master_file_bytes))

#     # Find the right sheet (first sheet or 'Sheet1')
#     sheet_names_upper = {s.upper(): s for s in wb.sheetnames}
#     if "SHEET1" in sheet_names_upper:
#         ws = wb[sheet_names_upper["SHEET1"]]
#     else:
#         ws = wb.active

#     # Find header row
#     header_row_idx = None
#     headers = {}  # col_letter/index → name
#     col_index_map = {}  # name → col_index (0-based)

#     for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
#         row_upper = [str(c).strip().upper() if c is not None else "" for c in row]
#         if "SHAPE" in row_upper:
#             header_row_idx = row_idx
#             for col_idx, val in enumerate(row_upper):
#                 if val:
#                     col_index_map[val] = col_idx
#             break

#     if header_row_idx is None:
#         raise ValueError("Could not find header row in Master File (looking for SHAPE column)")

#     def find_col(names):
#         for name in names:
#             if name in col_index_map:
#                 return col_index_map[name]
#         return None

#     shape_col     = find_col(["SHAPE"])
#     from_size_col = find_col(["FROM SIZE", "FROM_SIZE", "FROM"])
#     to_size_col   = find_col(["TO SIZE", "TO_SIZE", "TO"])
#     color_col     = find_col(["COLOR"])
#     clarity_col   = find_col(["CLARITY"])
#     grid_col      = find_col(["GRID"])
#     available_col = find_col(["AVAILABLE", "AVAILABL"])

#     if grid_col is None:
#         raise ValueError("Could not find 'Grid' column in Master File. Found: " + str(list(col_index_map.keys())))
#     if available_col is None:
#         raise ValueError("Could not find 'Available' column in Master File. Found: " + str(list(col_index_map.keys())))

#     updated_grid = 0
#     updated_available = 0
#     not_found_keys = []

#     all_rows = list(ws.iter_rows(min_row=header_row_idx + 1))

#     for row_cells in all_rows:
#         def cell_val(col_idx):
#             if col_idx is None or col_idx >= len(row_cells):
#                 return None
#             return row_cells[col_idx].value

#         shape   = cell_val(shape_col)
#         if not shape or str(shape).strip().upper() in ["", "SHAPE", "TOTAL"]:
#             continue

#         color     = cell_val(color_col)
#         clarity   = cell_val(clarity_col)
#         from_size = cell_val(from_size_col)
#         to_size   = cell_val(to_size_col)

#         key = make_key(shape, color, clarity, from_size, to_size)

#         # ---- GRID column: first try Grid Report INHAND, then Master Stock MAX PCS ----
#         grid_value = 0

#         # Priority 1: Grid Report INHAND
#         if key in inhand_map_grid:
#             grid_value = inhand_map_grid[key]

#         # Priority 2: Master Stock MAX PCS (overrides/fills if grid report had no match)
#         if key in master_stock_map:
#             max_pcs = master_stock_map[key]["max_pcs"]
#             # Use master_stock max_pcs for grid (as per requirement: copy Max PCS → Grid)
#             grid_value = max_pcs
#         elif key not in inhand_map_grid:
#             not_found_keys.append(key)

#         # Write Grid value (fill blank with 0)
#         grid_cell = row_cells[grid_col]
#         grid_cell.value = int(grid_value) if grid_value == int(grid_value) else grid_value
#         if grid_cell.value == 0 or grid_cell.value is None:
#             grid_cell.value = 0
#         updated_grid += 1

#         # ---- AVAILABLE column: Master Stock INHAND ----
#         available_value = 0
#         if key in master_stock_map:
#             available_value = master_stock_map[key]["inhand"]

#         avail_cell = row_cells[available_col]
#         avail_cell.value = int(available_value) if available_value == int(available_value) else available_value
#         updated_available += 1

#     output = BytesIO()
#     wb.save(output)
#     output.seek(0)
#     return output.getvalue(), updated_grid, updated_available, not_found_keys


# # ================================================================
# # ----------------------- UI LAYOUT ------------------------------
# # ================================================================

# col1, col2, col3 = st.columns(3)

# with col1:
#     st.markdown("""
#     <div class="upload-box">
#         <div class="step-badge">STEP 1</div>
#         <div class="upload-title">📊 Grid Report File</div>
#         <div class="upload-desc">Source of INHAND_PCS data — will be mapped to Master File's <b>Grid</b> column</div>
#     </div>
#     """, unsafe_allow_html=True)
#     grid_file = st.file_uploader("Upload Grid Report (.xlsx)", type=["xlsx", "xls"], key="grid")

# with col2:
#     st.markdown("""
#     <div class="upload-box">
#         <div class="step-badge">STEP 2</div>
#         <div class="upload-title">📦 Master Stock File</div>
#         <div class="upload-desc">Source of <b>Max PCS</b> → Grid column &amp; <b>INHAND</b> → Available column</div>
#     </div>
#     """, unsafe_allow_html=True)
#     master_stock_file = st.file_uploader("Upload Master Stock (.xlsx)", type=["xlsx", "xls"], key="master_stock")

# with col3:
#     st.markdown("""
#     <div class="upload-box">
#         <div class="step-badge">STEP 3</div>
#         <div class="upload-title">🎯 Master File</div>
#         <div class="upload-desc">Output file — <b>Grid</b> &amp; <b>Available</b> columns will be updated</div>
#     </div>
#     """, unsafe_allow_html=True)
#     master_file = st.file_uploader("Upload Master File (.xlsx)", type=["xlsx", "xls"], key="master")

# st.markdown("<hr>", unsafe_allow_html=True)

# # ---- Process Button ----
# if st.button("⚡ Run DiamondFlow Automation"):
#     if not grid_file or not master_stock_file or not master_file:
#         st.markdown('<div class="warn-box">⚠️ Please upload all 3 files before running.</div>', unsafe_allow_html=True)
#     else:
#         try:
#             with st.spinner("Processing diamond data..."):

#                 # Step 1: Read INHAND from Grid Report
#                 st.markdown('<div class="info-box">🔍 Step 1: Reading INHAND data from Grid Report...</div>', unsafe_allow_html=True)
#                 inhand_map = read_grid_inhand(grid_file)
#                 st.markdown(f'<div class="success-box">✅ Grid Report: Found <b>{len(inhand_map)}</b> INHAND records</div>', unsafe_allow_html=True)

#                 # Step 2: Read MAX PCS + INHAND from Master Stock
#                 st.markdown('<div class="info-box">🔍 Step 2: Reading Max PCS &amp; INHAND from Master Stock (MASTER sheet)...</div>', unsafe_allow_html=True)
#                 master_stock_map = read_master_stock(master_stock_file)
#                 st.markdown(f'<div class="success-box">✅ Master Stock: Found <b>{len(master_stock_map)}</b> records</div>', unsafe_allow_html=True)

#                 # Step 3: Update Master File
#                 st.markdown('<div class="info-box">🔍 Step 3: Updating Grid &amp; Available columns in Master File...</div>', unsafe_allow_html=True)
#                 master_bytes = master_file.read()
#                 output_bytes, grid_count, avail_count, not_found = update_master_file(
#                     master_bytes, inhand_map, master_stock_map
#                 )

#                 st.markdown(f"""
#                 <div class="success-box">
#                     ✅ Master File updated!<br>
#                     &nbsp;&nbsp;• <b>Grid</b> column: {grid_count} rows updated (blanks filled with 0)<br>
#                     &nbsp;&nbsp;• <b>Available</b> column: {avail_count} rows updated
#                 </div>
#                 """, unsafe_allow_html=True)

#                 if not_found:
#                     st.markdown(f'<div class="warn-box">⚠️ {len(not_found)} rows in Master File had no matching key in either Grid Report or Master Stock — filled with 0.</div>', unsafe_allow_html=True)

#             # ---- Summary Metrics ----
#             st.markdown("<hr>", unsafe_allow_html=True)
#             m1, m2, m3, m4 = st.columns(4)
#             m1.metric("Grid Report Records", len(inhand_map))
#             m2.metric("Master Stock Records", len(master_stock_map))
#             m3.metric("Grid Rows Updated", grid_count)
#             m4.metric("Available Rows Updated", avail_count)

#             # ---- Download ----
#             st.markdown("<br>", unsafe_allow_html=True)
#             st.download_button(
#                 label="⬇️ Download Updated Master File",
#                 data=output_bytes,
#                 file_name="Master_File_Updated.xlsx",
#                 mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#             )

#         except Exception as e:
#             st.markdown(f'<div class="warn-box">❌ Error: {str(e)}</div>', unsafe_allow_html=True)
#             st.exception(e)

# # ---- Footer info ----
# st.markdown("<hr>", unsafe_allow_html=True)
# st.markdown("""
# <div style="text-align:center; color:#666; font-size:12px; margin-top:10px;">
#     <b>DiamondFlow AI</b> — Automation Logic:<br>
#     Grid Report <code>INHAND_PCS</code> → Master File <b>Grid</b> column &nbsp;|&nbsp;
#     Master Stock <code>Max PCS</code> → Master File <b>Grid</b> column &nbsp;|&nbsp;
#     Master Stock <code>INHAND</code> → Master File <b>Available</b> column &nbsp;|&nbsp;
#     All blanks filled with <b>0</b>
# </div>
# """, unsafe_allow_html=True)





# app.py
# ==========================================================
# FIXED VERSION
# Copy/Paste VALUES ONLY
# No formulas inserted
# No #REF issue
# Download all 3 files
# # ==========================================================

import streamlit as st
from openpyxl import load_workbook
from io import BytesIO

st.set_page_config(page_title="Daily Work Automation", layout="wide")
st.title("📊 Full Daily Work Automation")

# --------------------------------------------------
# SESSION
# --------------------------------------------------
if "processed" not in st.session_state:
    st.session_state.processed = False

# --------------------------------------------------
# SHAPES
# --------------------------------------------------
SHAPES = [
    "ROUND","OVAL","EMERALD","RADIANT","PEAR",
    "PRINCESS","MARQUISE","CUSHION MODIFIED",
    "CUSHION BRILLIANT","ASSCHER","HEART"
]

# --------------------------------------------------
# HELPERS
# --------------------------------------------------
def clean(v):
    if v is None:
        return ""
    return str(v).strip().upper()

def is_number(v):
    try:
        float(v)
        return True
    except:
        return False

def find_header(ws, text):
    for row in ws.iter_rows():
        for cell in row:
            if clean(cell.value) == clean(text):
                return cell.row, cell.column
    return None, None

def find_shape_rows(ws):
    found = {}

    for r in range(1, ws.max_row + 1):
        val = clean(ws.cell(r,1).value)

        if val in SHAPES and val not in found:
            found[val] = r

    return found

def find_total_row(ws, start):
    for r in range(start, ws.max_row + 1):
        for c in range(1,10):
            if clean(ws.cell(r,c).value) == "TOTAL":
                return r
    return start + 15

# --------------------------------------------------
# STEP 1 GRID -> MASTER STOCK
# --------------------------------------------------
def copy_grid_to_stock(ws_grid, ws_stock):

    g = find_shape_rows(ws_grid)
    s = find_shape_rows(ws_stock)

    for shape in SHAPES:

        if shape in g and shape in s:

            gs = g[shape]
            ge = find_total_row(ws_grid, gs)

            ss = s[shape]

            rows = ge - gs

            for i in range(rows):

                for col in range(1, ws_grid.max_column + 1):

                    val = ws_grid.cell(gs+i, col).value

                    if is_number(val):
                        ws_stock.cell(ss+i, col).value = val

# --------------------------------------------------
# COPY VALUE ONLY
# --------------------------------------------------
def copy_value_column(ws_source, source_header, ws_target, target_header):

    sr, sc = find_header(ws_source, source_header)
    tr, tc = find_header(ws_target, target_header)

    if not sr or not tr:
        return

    r = sr + 1
    t = tr + 1

    while r <= ws_source.max_row:

        val = ws_source.cell(r, sc).value

        if r > sr + 1000:
            break

        # IMPORTANT: copy only final values
        if val is None or clean(val) == "":
            ws_target.cell(t, tc).value = 0
        else:
            ws_target.cell(t, tc).value = val

        r += 1
        t += 1

# --------------------------------------------------
# PROCESS
# --------------------------------------------------
def process(grid_file, stock_file, master_file):

    # normal files
    wb_grid = load_workbook(grid_file)
    wb_stock = load_workbook(stock_file)
    wb_master = load_workbook(master_file)

    ws_grid = wb_grid.active
    ws_stock = wb_stock.active
    ws_master = wb_master.active

    # ------------------------------------------------
    # STEP 1
    # Grid -> Master Stock
    # ------------------------------------------------
    copy_grid_to_stock(ws_grid, ws_stock)

    # Save updated stock first
    temp = BytesIO()
    wb_stock.save(temp)
    temp.seek(0)

    # ------------------------------------------------
    # IMPORTANT FIX
    # Load SAME updated file values
    # ------------------------------------------------
    wb_stock_values = load_workbook(temp, data_only=False)

    # Sheet2
    ws_sheet2 = wb_stock_values.worksheets[1]

    # ------------------------------------------------
    # STEP 2
    # INHAND -> Available
    # MAX PCS -> Grid
    # ------------------------------------------------
    copy_value_column(ws_sheet2, "INHAND", ws_master, "AVAILABLE")
    copy_value_column(ws_sheet2, "MAX PCS", ws_grid, "GRID")

    # ------------------------------------------------
    # SAVE
    # ------------------------------------------------
    out1 = BytesIO()
    out2 = BytesIO()
    out3 = BytesIO()

    wb_grid.save(out1)
    wb_stock.save(out2)
    wb_master.save(out3)

    out1.seek(0)
    out2.seek(0)
    out3.seek(0)

    return out1, out2, out3

# --------------------------------------------------
# UI
# --------------------------------------------------
grid_file = st.file_uploader("Upload Grid Report", type=["xlsx"])
stock_file = st.file_uploader("Upload Master Stock File", type=["xlsx"])
master_file = st.file_uploader("Upload Master File", type=["xlsx"])

if grid_file and stock_file and master_file:

    if st.button("🚀 Process All Files"):

        try:
            with st.spinner("Processing..."):

                g, s, m = process(grid_file, stock_file, master_file)

                st.session_state.grid = g.getvalue()
                st.session_state.stock = s.getvalue()
                st.session_state.master = m.getvalue()

                st.session_state.processed = True

            st.success("✅ Completed Successfully")

        except Exception as e:
            st.error(str(e))

# --------------------------------------------------
# DOWNLOADS
# --------------------------------------------------
if st.session_state.processed:

    st.download_button(
        "📥 Download Grid Report",
        data=st.session_state.grid,
        file_name="Updated_Grid_Report.xlsx"
    )

    st.download_button(
        "📥 Download Master Stock",
        data=st.session_state.stock,
        file_name="Updated_Master_Stock.xlsx"
    )

    st.download_button(
        "📥 Download Master File",
        data=st.session_state.master,
        file_name="Updated_Master_File.xlsx"
    )

